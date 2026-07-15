#!/usr/bin/env python3
"""Extract static game data from the W&R Soviet Republic planning workbook to JSON."""
import json, re, os
import openpyxl

SRC = 'workers.xlsx'
OUT = 'data'
os.makedirs(OUT, exist_ok=True)

wbF = openpyxl.load_workbook(SRC, data_only=False)
wbV = openpyxl.load_workbook(SRC, data_only=True)

# ---------- translations ----------
wsF, wsV = wbF['Übersetzung'], wbV['Übersetzung']
T = {}
for row in wsV.iter_rows(min_row=3):
    rid = row[0].value
    if rid is None:
        continue
    rid = int(rid)
    de, en = row[1].value, row[2].value
    T[rid] = {'de': de if de is not None else '', 'en': en if en is not None else ''}

FALLBACK_RE = re.compile(r'^\s*please select language\s*$', re.I)
TOKEN_RE = re.compile(r'VLOOKUP\((\d+)\s*,|"((?:[^"]|"")*)"')

def resolve_name(formula, lang):
    """Best-effort: resolve a name formula (VLOOKUPs into Übersetzung + literals) to text."""
    parts = []
    for m in TOKEN_RE.finditer(formula):
        if m.group(1) is not None:
            rid = int(m.group(1))
            if rid in T:
                parts.append(T[rid][lang])
        else:
            lit = m.group(2).replace('""', '"')
            if FALLBACK_RE.match(lit):
                continue
            parts.append(lit)
    return ''.join(parts).strip()

# fixes for machine-translated EN terms in the sheet's translation table
EN_FIXES = {'Eat': 'Food'}

def cell_names(cF, cV):
    """Return (de, en) for a cell that is either a literal or a translation formula."""
    f = cF.value
    v = cV.value
    if f is None:
        return None, None
    if isinstance(f, str) and f.startswith('='):
        de = resolve_name(f, 'de')
        en = resolve_name(f, 'en')
        # trust the cached (German) value if resolution failed
        if not de and isinstance(v, str):
            de = v
        if not en:
            en = de
        return de, EN_FIXES.get(en, en)
    return str(f), str(f)

def num(x, default=0):
    if x is None or x == '' or x == '-':
        return default
    try:
        return round(float(x), 6)
    except (TypeError, ValueError):
        return default

# ---------- resources (key <-> translation id) + default prices ----------
wsV_i = wbV['Importdata Preise']
resources = []
seen = set()
for r in range(3, 59):
    tid = wsV_i.cell(row=r, column=1).value
    key = wsV_i.cell(row=r, column=2).value
    if key is None:
        continue
    entry = {'key': key}
    if tid is not None:
        tid = int(tid)
        entry['tid'] = tid
        entry['de'] = T.get(tid, {}).get('de', key)
        entry['en'] = T.get(tid, {}).get('en', key)
    else:
        entry['de'] = key
        entry['en'] = key
    if key not in seen:
        resources.append(entry)
        seen.add(key)
# workers pseudo-resource (id 37 = Arbeiter)
resources.append({'key': 'workers', 'tid': 37, 'de': T[37]['de'], 'en': T[37]['en']})

def read_price_block(start):
    d = {}
    for r in range(start, start + 56):
        key = wsV_i.cell(row=r, column=2).value
        val = wsV_i.cell(row=r, column=3).value
        if key and key != '$end':
            d[key] = num(val)
    return d

default_prices = {
    'purchaseUSD': read_price_block(3),
    'purchaseRUB': read_price_block(63),
    'sellUSD': read_price_block(123),
    'sellRUB': read_price_block(183),
    'deliveryCostUSD': num(wsV_i.cell(row=361, column=3).value),
    'deliveryCostRUB': num(wsV_i.cell(row=362, column=3).value),
    'workdayCostUSD': num(wsV_i.cell(row=363, column=3).value),
    'workdayCostRUB': num(wsV_i.cell(row=364, column=3).value),
    'imigrantCostRUB': num(wsV_i.cell(row=365, column=3).value),
    'imigrantCostUSD': num(wsV_i.cell(row=366, column=3).value),
}
json.dump({'resources': resources, 'defaults': default_prices},
          open(f'{OUT}/resources.json', 'w'), ensure_ascii=False, indent=1)
print('resources:', len(resources))

# ---------- production buildings (Datenblatt Produktion) ----------
wsF_p, wsV_p = wbF['Datenblatt Produktion'], wbV['Datenblatt Produktion']
prod = []
for r in range(2, wsV_p.max_row + 1):
    if wsF_p.cell(row=r, column=2).value is None:
        continue
    g_de, g_en = cell_names(wsF_p.cell(row=r, column=1), wsV_p.cell(row=r, column=1))
    n_de, n_en = cell_names(wsF_p.cell(row=r, column=2), wsV_p.cell(row=r, column=2))
    if not n_de:
        continue
    def rn(col):  # resource name pair at col
        return cell_names(wsF_p.cell(row=r, column=col), wsV_p.cell(row=r, column=col))
    def v(col):
        return num(wsV_p.cell(row=r, column=col).value)
    prods, cons = [], []
    for c in (4, 6):  # D,F
        nm_de, nm_en = rn(c)
        rate = v(c + 1)
        if nm_de and rate:
            prods.append({'de': nm_de, 'en': nm_en, 'rate': rate})
    for c in (8, 10, 12, 14, 16):  # H..P
        nm_de, nm_en = rn(c)
        rate = v(c + 1)
        if nm_de and rate:
            cons.append({'de': nm_de, 'en': nm_en, 'rate': rate})
    prod.append({
        'group': {'de': g_de, 'en': g_en},
        'de': n_de, 'en': n_en,
        'workers': v(3),
        'production': prods,
        'consumption': cons,
        'power': v(18),        # R Stromverbrauch (MWh/day)
        'maxKW': v(19),        # S Max Wattzahl
        'water': v(20),        # T
        'hotwater': v(21),     # U
        'wastePerWorker': v(23),  # W
        'workdays': v(24),     # X
        'gravel': v(25), 'bricks': v(26), 'steel': v(27), 'concrete': v(28),
        'asphalt': v(29), 'boards': v(30), 'panels': v(31),
        'ecomponents': v(32), 'mcomponents': v(33),
    })
json.dump(prod, open(f'{OUT}/production_buildings.json', 'w'), ensure_ascii=False, indent=1)
print('production buildings:', len(prod))

# ---------- city buildings (Daten StadtData City) ----------
wsF_c, wsV_c = wbF['Daten StadtData City'], wbV['Daten StadtData City']
city = []
for r in range(2, wsV_c.max_row + 1):
    if wsF_c.cell(row=r, column=1).value is None:
        continue
    n_de, n_en = cell_names(wsF_c.cell(row=r, column=1), wsV_c.cell(row=r, column=1))
    t_de, t_en = cell_names(wsF_c.cell(row=r, column=2), wsV_c.cell(row=r, column=2))
    if not n_de:
        continue
    def v(col, default=0):
        return num(wsV_c.cell(row=r, column=col).value, default)
    kind = wsV_c.cell(row=r, column=3).value or ''
    quality = wsV_c.cell(row=r, column=4).value
    quality = num(quality, None) if quality not in (None, '-') else None
    city.append({
        'de': n_de, 'en': n_en,
        'type': {'de': t_de, 'en': t_en},
        'kind': kind,
        'quality': quality,
        'workers': v(5),
        'special': v(6),
        'visitors': v(7),
        'inhabitants': v(8),
        'power': v(9),
        'maxKW': v(10),
        'water': v(11),
        'hotwater': v(12),
        'waste': v(13),
        'workdays': v(16),
        'gravel': v(17), 'bricks': v(18), 'steel': v(19), 'concrete': v(20),
        'asphalt': v(21), 'boards': v(22), 'panels': v(23),
        'ecomponents': v(24), 'mcomponents': v(25),
        'recommendedFor': v(27),
    })
json.dump(city, open(f'{OUT}/city_buildings.json', 'w'), ensure_ascii=False, indent=1)
print('city buildings:', len(city))

# ---------- vehicles (Datenblatt Fahrzeuge) ----------
wsF_v, wsV_v = wbF['Datenblatt Fahrzeuge'], wbV['Datenblatt Fahrzeuge']
# header row 2: resolve names
headers = {1: {'de': 'Typ', 'en': 'Type'}}
for c in range(2, wsV_v.max_column + 1):
    h_de, h_en = cell_names(wsF_v.cell(row=2, column=c), wsV_v.cell(row=2, column=c))
    if h_de:
        headers[c] = {'de': h_de, 'en': h_en}
vehicles = []
for r in range(3, wsV_v.max_row + 1):
    name = wsV_v.cell(row=r, column=2).value
    if not name:
        continue
    row = {}
    for c, h in headers.items():
        val = wsV_v.cell(row=r, column=c).value
        if val is None or c == 2:
            continue
        if isinstance(val, (int, float)):
            row[h['de']] = round(float(val), 4)
        else:
            row[h['de']] = val
    vehicles.append({'name': str(name), 'attrs': row})
json.dump({'headers': [h for h in headers.values()], 'vehicles': vehicles},
          open(f'{OUT}/vehicles.json', 'w'), ensure_ascii=False, indent=1)
print('vehicles:', len(vehicles))

# ---------- decade price presets (Preisanstiege) ----------
wsV_a = wbV['Preisanstiege']
decades = {}
# blocks: (year, ware_col, sellRUB, buyRUB, sellUSD, buyUSD) 1-indexed
blocks = [(1920, 1, 2, 3, 4, 5), (1930, 6, 7, 9, 11, 12), (1940, 13, 14, 16, 18, 20),
          (1950, 22, 23, 25, 27, 29), (1960, 31, 32, 34, 36, 38),
          (1970, 40, 41, 42, 43, 44), (1980, 45, 46, 47, 48, 49)]
# map German ware name -> resource key
de2key = {r['de']: r['key'] for r in resources}
for year, wc, srb, brb, susd, busd in blocks:
    d = {}
    for r in range(3, 58):
        ware = wsV_a.cell(row=r, column=wc).value
        if not ware:
            continue
        key = de2key.get(str(ware).strip())
        if not key:
            continue
        d[key] = {'sellRUB': num(wsV_a.cell(row=r, column=srb).value),
                  'buyRUB': num(wsV_a.cell(row=r, column=brb).value),
                  'sellUSD': num(wsV_a.cell(row=r, column=susd).value),
                  'buyUSD': num(wsV_a.cell(row=r, column=busd).value)}
    decades[year] = d
json.dump(decades, open(f'{OUT}/decade_prices.json', 'w'), ensure_ascii=False, indent=1)
print('decades:', {y: len(v) for y, v in decades.items()})

# ---------- UI translations (de/en) ----------
json.dump(T, open(f'{OUT}/translations.json', 'w'), ensure_ascii=False, indent=1)
print('translations:', len(T))
